import os
import pandas as pd
import xlrd
import openpyxl

def handle_excel_xls(file_path):
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)

    # Column: Mã hàng, Tên hàng, Tồn cuối
    column_indices = [1, 4, 22]
    products = []
    for row_index in range(sheet.nrows):
        if row_index > 8:
            row_values = [sheet.cell_value(row_index, col_index) for col_index in column_indices]
            if row_values[0] != '' and row_values[1] != '' and row_values[2] != '':
                value_column = {
                    "product_code": row_values[0],
                    "product_name": row_values[1],
                    "ending_inventory": row_values[2],
                }
                products.append(value_column)
                print(f"Row {row_index + 1}: {value_column}")
    return products


def handle_excel_xlsx(excel_file_path):
    workbook = openpyxl.load_workbook(excel_file_path)
    sheet = workbook['Sheet1']

    # Column name: China, Vietnam
    column_indices = [1, 2] 

    product_names = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = [row[i - 1] for i in column_indices]
        if values[1] is not None and values[1] != '' and values[1] != '越南型号':
            product_names.append(values[1].lower().strip())
        print("Row {}: Cell values: {}".format(row_index, values))
    
    # Close the workbook
    workbook.close()
    return product_names


def get_strings_with_same_word_count(product_stocks, product_name):
    found_product_names = []
    for product_stock in product_stocks:
        if product_stock != None:
            count_equal_elements = compare_strings_count_equal_elements(product_stock, product_name)
            if count_equal_elements:
                found_product_names.append({"is_found": True, "product_name": product_name, "product_stock_name": product_stock})
            else:
                found_product_names.append({"is_found": False, "product_name": product_name, "product_stock_name": product_stock})
    return found_product_names


def compare_strings_count_equal_elements(product_name, product_stock_name):
    words1 = set(product_stock_name.lower().strip().split())
    words2 = set(product_name.lower().strip().split())
    count_equal_elements = sum(1 for elem1 in words1 if elem1 in words2)
    
    # Count word equal True
    if product_name.lower().strip() in product_stock_name.lower().strip():
        return True
    elif count_equal_elements >= 4:
        return True
    else:
        return False


def remove_files_folder(folder_path): 
    files = os.listdir(folder_path)
    for file in files:
        file_path = os.path.join(folder_path, file)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
                print(f"File {file_path} removed successfully.")
            else:
                print(f"{file_path} is not a file.")
        except Exception as e:
            print(f"Error removing {file_path}: {e}")
            
            
def filter_products(data_stocks, data_products):
    products = []
    product_stock_names = []
    for product in data_products:
        founds = get_strings_with_same_word_count(data_stocks, product['product_name'].lower().strip())
        for found in founds:
            if found['is_found'] == True and product['ending_inventory'] < 10:
                products.append({
                    'product_code': product['product_code'], 
                    'product_name': product['product_name'].lower().strip(),
                    'ending_inventory': product['ending_inventory'],
                    'need_to_get': 10,
                })
                product_stock_names.append({
                    'product_code': product['product_code'], 
                    'product_name': product['product_name'].lower().strip(), 
                    'product_stock_name': found['product_stock_name'].lower().strip(), 
                    'ending_inventory': product['ending_inventory'],
                    'need_to_get': 10,
                })
    
    product_filters = remove_duplicates(products)
    product_stock_names = remove_duplicates(product_stock_names)
    return {
        "product_filters": product_filters, 
        "product_stock_names": product_stock_names, 
        "data_stocks": list(dict.fromkeys(data_stocks))
    }
    

def filter_product_stocks(data_stocks, data_products):
    products = []
    product_stock_names = []
    for pro_stock_name in data_stocks:
        found_products = get_product_compare(data_products, pro_stock_name.lower().strip())
        product_stock_names.extend(found_products)
        for found_product in found_products:
            products.append({
                'product_code': found_product['product_code'], 
                'product_name': found_product['product_name'].lower().strip(),
                'ending_inventory': found_product['ending_inventory'],
                'need_to_get': found_product['need_to_get'],
            })
            
    product_filters = remove_duplicates(products)
    product_stock_names = remove_duplicates(product_stock_names)
    return {
        "product_filters": product_filters, 
        "product_stock_names": product_stock_names, 
        "data_stocks": list(dict.fromkeys(data_stocks))
    }


def get_product_compare(data_products, product_stock_name):
    found_products = []
    for product in data_products:
        if product != None:
            is_compare_equal = compare_strings_count_equal_elements(product['product_name'], product_stock_name)
            if is_compare_equal and product['ending_inventory'] < 10:
                found_products.append({
                    'product_code': product['product_code'], 
                    'product_name': product['product_name'].lower().strip(), 
                    'product_stock_name': product_stock_name.lower().strip(), 
                    'ending_inventory': product['ending_inventory'],
                    'need_to_get': 10,
                })
    return found_products


def remove_duplicates(data_list):
    # Check for duplicates
    seen_product_codes = set()
    unique_data_list = []

    for data_dict in data_list:
        product_code = data_dict["product_code"]
        
        # Check if product_code is already seen
        if product_code not in seen_product_codes:
            seen_product_codes.add(product_code)
            unique_data_list.append(data_dict)

    # Print the unique list
    return unique_data_list


def save_file_new(data):
    # Create a new Excel workbook
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Replace names in header_titles with translations
    header_titles = ["Mã sản phẩm", "Tên sản phẩm", "Tồn cuối", "Cần lấy"]

    # Add header titles to the first row
    for col_num, title in enumerate(header_titles, start=1):
        sheet.cell(row=1, column=col_num, value=title)

    # Add data to the sheet
    for row_num, row_data in enumerate(data, start=2):
        for col_num, key in enumerate(["product_code", "product_name", "ending_inventory", "need_to_get"], start=1):
            sheet.cell(row=row_num, column=col_num, value=row_data[key])

    # Set column width to fit content
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        try:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        except:
            pass

        # Save the workbook to a file
        workbook.save('static/excels/LocHangTon.xlsx')